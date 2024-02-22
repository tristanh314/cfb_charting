# Import Dependencies

import numpy as np
import pandas as pd
from bs4 import BeautifulSoup
import re
import openpyxl as op
import sys

def find_off(drive_table):
    """
    Input: A BeautifulSoup object corresponding to a table of play information for a single drive from CBS cfb play-by-play.
    Output: The abbreviation of the team that is on offense during the drive in question.
    """
    team_tag = drive_table.find('a', class_='')
    team_link = team_tag.get('href')
    team_split = team_link.split('/')
    team_name = team_split[3]
    return team_name


def parse_play(play, team):
    """
    Input: A dataframe of plays for a single team, eg, output of scrape_cbs.
    Output: A dataframe of values that can be written to an ATQ charting template.
    """

    # Create a series to store information for the output row.
    row = pd.Series({'E':'', 'F':np.NaN, 'G':'', 'AQ':np.NaN, 'AR':np.NaN, 'AS':np.NaN, 'AT':np.NaN, 'AV':np.NaN, 'AW':'', 'AX':''})

    # Determine yardage gained/lost if appropriate and and write to the output row.
    if  (re.compile('\+').search(play['Result'])) != None:
        row['AV'] = int(re.compile('[0-9]*\s').search(play['Result']).group())
    elif  (re.compile('\-').search(play['Result'])) != None:
        row['AV'] = -int(re.compile('[0-9]*\s').search(play['Result']).group())
    elif (re.compile('No Gain').search(play['Result'])) != None:
        row['AV'] = 0
    elif (re.compile('Int').search(play['Result'])) != None:
        row['AV'] = 0
    elif (re.compile('Penalty').search(play['Result'])) != None:
        row['AV'] = np.NaN
        row['AW'] = 'Penalty'
    elif (re.compile('Sack').search(play['Result'])) != None:
        row['AV'] = int(re.compile('for \-[0-9]*').search(play['Description']).group()[4:])
    
    # Determine the play type.
    if (re.compile('pass complete').search(play['Description'])) != None:
        row['G'] = 'p'
        if (re.compile('TOUCHDOWN').search(play['Description'])) != None:
            row['AW'] = 'Passing Touchdown'
        else:
            row['AW'] = 'Pass Reception'
    elif (re.compile('pass incomplete').search(play['Description'])) != None:
        row['G'] = 'p'
        row['AW'] = 'Pass Incompletion'
    elif (re.compile('scrambles').search(play['Description'])) != None:
        row['G'] = 'p'
        if (re.compile('TOUCHDOWN').search(play['Description'])) != None:
            row['AW'] = 'Rushing Touchdown'
        else:
            row['AW'] = 'Rush'
    elif (re.compile('rushed').search(play['Description'])) != None:
        row['G'] = 'r'
        if (re.compile('TOUCHDOWN').search(play['Description'])) != None:
            row['AW'] = 'Rushing Touchdown'
        else:
            row['AW'] = 'Rush'
    else:
        row['G'] = '?'
    
    # Output the time stamp in the appropriate columns.
    timestamp = re.compile('[0-9]*:[0-9][0-9]\s\-\s[0-9]').search(play['Description']).group()
    min_sec_qtr = re.compile('[0-9]+').findall(timestamp)
    row['AQ'] = int(min_sec_qtr[2])
    row['AR'] = int(min_sec_qtr[0])
    if min_sec_qtr[1][0] == '0':
        row['AS'] = int(min_sec_qtr[1][-1])
    else:
        row['AS'] = int(min_sec_qtr[1])

    # extract, distance, and field position.
    down_dist_fp_str = re.compile('.*[(]').match(play['Description']).group()
    down_dist_fp = re.compile('[0-9][0-9]?').findall(down_dist_fp_str)
    
    # Write the down in the proper output format.
    try:
        down = down_dist_fp[0]
    except:
        down = '?'
    if down == '1':
        row['E'] = '1st'
    elif down == '2':
        row['E'] = '2nd'
    elif down == '3':
        row['E'] = '3rd'
    elif down == '4':
        row['E'] = '4th'
    else:
        row['E'] = '?'

    # Write the distance in the proper output format.
    try:
        dist = int(down_dist_fp[1])
    except:
        dist = 0
    row['F'] = dist

    # Write the field position in the proper output format.
    try:
        fp = int(down_dist_fp[2])
    except:
        fp = 0
    if re.compile(team).search(down_dist_fp_str) == None:
        row['AT'] = fp
    else:
        row['AT'] = 100-fp

    # Include full play description for reference.
    row['AX'] = play['Description']

    return row

def write_drive(drive_df, team, ws, t_row):
    """
    Input: A dataframe containing drive information formatted to be written into the ATQ charting template, and a worksheet from an open Openpyxl workbook, row number in the worksheet to begin writing to.
    Output: The number of rows written to the worksheet. The play information is written to a template .xlsx file.
    """
    # Remove special teams plays
    for index in range(0,drive_df.shape[0]):
        if re.compile('kicks').search(drive_df.loc[index]['Description']) != None:
            drive_df.drop(index,inplace=True)
        elif re.compile('extra\spoint').search(drive_df.loc[index]['Description']) != None:
            drive_df.drop(index,inplace=True)
        elif re.compile('punts').search(drive_df.loc[index]['Description']) != None:
            drive_df.drop(index,inplace=True)
        elif re.compile('field\sgoal').search(drive_df.loc[index]['Description']) != None:
            drive_df.drop(index,inplace=True)
        else:
            pass
    
    # Reindex the dataframe
    drive_df.reset_index(inplace=True)
    drive_df.drop(columns=['index'])
        
    # Write information for each play to the worksheet.
    index = 0
    while index < drive_df.shape[0]:
        play_info = parse_play(drive_df.iloc[index], team)
        ws[f'E{t_row}']=play_info['E']
        ws[f'F{t_row}']=play_info['F']
        ws[f'G{t_row}']=play_info['G']
        ws[f'AQ{t_row}']=play_info['AQ']
        ws[f'AR{t_row}']=play_info['AR']
        ws[f'AS{t_row}']=play_info['AS']
        ws[f'AT{t_row}']=play_info['AT']
        ws[f'AV{t_row}']=play_info['AV']
        ws[f'AW{t_row}']=play_info['AW']
        ws[f'AX{t_row}']=play_info['AX']
        index+=1
        t_row+=1

    return drive_df.shape[0]

def main(html_file, xlsx_file):
    """
    Input: Relative file path for donwloaded html of college football plays from a game recorded at cbssports.com.
    Output: The home and away teams. The individual play data is written to a .xlsx file. 
    """

    # Store html from the downloaded page as a soup object.
    page=open(html_file)
    plays_soup = BeautifulSoup(page.read(), 'html.parser')
    page.close()

    # Determine the names of the teams in the game.
    home_div = plays_soup.find('div', class_='hud-table-cell team-name-container full home')
    home_abbr = home_div.find('div', class_='abbr').text
    home =  re.compile('[A-Z]+').search(home_abbr).group() 
    away_div = plays_soup.find('div', class_='hud-table-cell team-name-container full away')
    away_abbr = away_div.find('div', class_='abbr').text
    away =  re.compile('[A-Z]+').search(away_abbr).group()

    # Collect a list of drive tables.
    drive_set = plays_soup.find_all('div', id='TableBase')

    # Read the information from html drive tables into pandas dataframes.
    drive_list = []

    for drive in range(0, len(drive_set)):
        drive_df = pd.DataFrame({'Offense':[],'Result':[],'Description':[]})
        drive_off = find_off(drive_set[drive])
        drive_table = drive_set[drive].find('tbody')
        drive_plays = drive_table.find_all('tr')
        for index in range(0, len(drive_plays)):
            play_row = drive_plays[index].find_all('td')
            play_result = play_row[0].text
            play_desc = play_row[1].text
            drive_df.loc[len(drive_df.index)] = pd.Series({'Offense':drive_off, 'Result':play_result, 'Description':play_desc})
        drive_list += [drive_df]

    # Split the drives into two lists, one for each team.
    home_drive_list = [drive for drive in drive_list if drive['Offense'][0] == home]
    away_drive_list = [drive for drive in drive_list if drive['Offense'][0] == away]

    # Load the template
    wb = op.load_workbook(filename = 'atq_charting_template.xlsx')

    # Write the data for the home offense.
    home_ws = wb['PRIME Off, SEC Def']
    home_t_row = 1
    for drive_df in home_drive_list:
        home_t_row+=write_drive(drive_df, home, home_ws, home_t_row)+1

    # Write the data for the home offense.
    away_ws = wb['SEC Off, PRIME Def']
    away_t_row = 1
    for drive_df in away_drive_list:
        away_t_row+=write_drive(drive_df, away, away_ws, away_t_row)+1
        
        
    # Save the new .xlsx file
    wb.save(xlsx_file)
    
    return home, away

if __name__ == '__main__':
    main(sys.argv[1], 'output.xlsx')